# Cloud Function: fn_publish — перенос проверенных объектов на карту БЕЗ участия ноутбука.
# Облачный аналог 61_publish.py. Данные карты живут в бакете (префикс data/), поэтому функция
# самодостаточна: скачивает их, сливает принятые объекты, ДОСЧИТЫВАЕТ ТОЛЬКО НОВОЕ, пересобирает
# карту и выкладывает её. Полный пересчёт корпуса не делается — иначе каждая публикация
# перемалывала бы 2600+ объектов ради пары точек.
#
# Вызов: POST {} — опубликовать принятое; {"force": true} — пересобрать даже без новых объектов;
#        {"dry": true} — посчитать, но не выкладывать.
# ENV: BUCKET, SITE_BUCKET, AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, VERIFY_API (опц.)
import os, io, re, json, time, math, shutil, base64, runpy, collections
import urllib.request, urllib.parse
import boto3

BUCKET = os.environ.get("BUCKET", "loess-results")
SITE_BUCKET = os.environ.get("SITE_BUCKET", "loess-map")
DATA_PREFIX = "data/"
UPLOADS = "uploads/"
TMP = "/tmp/data"
MIN_KEEP = float(os.environ.get("MIN_KEEP", 0.9))     # карта не должна внезапно похудеть
UA = "loess-sections/1.0 (research)"

s3 = boto3.client("s3", endpoint_url="https://storage.yandexcloud.net", region_name="ru-central1",
                  aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
                  aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"])

def s3_get_json(key, default=None):
    try: return json.loads(s3.get_object(Bucket=BUCKET, Key=key)["Body"].read())
    except Exception: return default

def s3_put_json(key, obj):
    s3.put_object(Bucket=BUCKET, Key=key, Body=json.dumps(obj, ensure_ascii=False).encode("utf-8"),
                  ContentType="application/json")

def log(*a): print(*a, flush=True)

# ---------- 1. данные карты из бакета во временную папку ----------
DATA_FILES = ["records_clean_geo_v2.xlsx", "section_evidence.json", "dem_elevation.json",
              "accuracy_radius.json", "page_links_v2.json", "scan_index.json", "rotated_pages.json", "column_data.json", "object_kind.json",
              "merged_uploads.json", "template.html"]

def pull_data():
    os.makedirs(TMP, exist_ok=True)
    got = []
    for name in DATA_FILES:
        try:
            body = s3.get_object(Bucket=BUCKET, Key=DATA_PREFIX + name)["Body"].read()
            open(os.path.join(TMP, name), "wb").write(body); got.append(name)
        except Exception:
            pass
    log(f"[1] данные из бакета: {len(got)} файлов")
    return got

def push_data(names):
    ct = {".json": "application/json", ".html": "text/html; charset=utf-8",
          ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    for name in names:
        p = os.path.join(TMP, name)
        if os.path.exists(p):
            s3.put_object(Bucket=BUCKET, Key=DATA_PREFIX + name, Body=open(p, "rb").read(),
                          ContentType=ct.get(os.path.splitext(name)[1], "application/octet-stream"))

# ---------- 2. слияние принятых объектов (логика 58) ----------
def join(v):
    if isinstance(v, list): return "; ".join(str(x) for x in v if x not in (None, "", "ND"))
    return "" if v in (None, "ND") else str(v)

def thick(p, kind):
    out = []
    for t in p.get("thickness", []):
        m = re.match(r"\s*(\w+)\s*:\s*(.+?)\s*м?\s*$", str(t))
        if m and m.group(1) == kind: out.append(m.group(2))
    return "; ".join(out)

def elev(v):
    m = re.findall(r"-?\d+(?:[.,]\d+)?", str(v or ""))
    return (m[0].replace(",", "."), m[-1].replace(",", ".")) if m else ("", "")

def all_jobs():
    out, tok = [], None
    while True:
        kw = {"Bucket": BUCKET, "Prefix": UPLOADS, "MaxKeys": 1000}
        if tok: kw["ContinuationToken"] = tok
        r = s3.list_objects_v2(**kw)
        out += [o["Key"].split("/")[1] for o in r.get("Contents", []) if o["Key"].endswith("/status.json")]
        tok = r.get("NextContinuationToken")
        if not r.get("IsTruncated"): break
    return sorted(set(out))

def merge_accepted():
    """Дописывает принятые объекты в таблицу. Возвращает (сколько добавлено, какие публикации затронуты)."""
    import openpyxl
    xl = os.path.join(TMP, "records_clean_geo_v2.xlsx")
    ledger = {}
    lp = os.path.join(TMP, "merged_uploads.json")
    if os.path.exists(lp):
        try: ledger = json.load(open(lp, encoding="utf-8"))
        except Exception: ledger = {}

    wb = openpyxl.load_workbook(xl); ws = wb.active
    H = [c.value for c in ws[1]]; C = {n: i for i, n in enumerate(H)}
    # индекс существующих точек — чтобы объект того же места не создал дубль-маркер рядом
    existing = collections.defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        la, lo = row[C["lat"]], row[C["lon"]]
        if la in (None, "", "ND") or str(la) == "None": continue
        try: existing[str(row[C["nearest_locality"]] or "").strip().lower()].append((float(la), float(lo)))
        except Exception: pass

    all_pts = [p for pts in existing.values() for p in pts]

    def snap(loc, lat, lon, limit=300, any_limit=150):
        """Совмещаем с уже имеющейся точкой: сначала по совпадению названия, затем — по близости
        с любой точкой. Второе нужно из-за разного написания («Отказное» и «Otkaznoe» — одно место,
        а для сравнения строк это разные названия)."""
        best, bd = None, limit
        for la, lo in existing.get(str(loc or "").strip().lower(), []):
            d = math.hypot((la - lat) * 111320, (lo - lon) * 111320 * math.cos(math.radians(lat)))
            if d < bd: best, bd = (la, lo), d
        if best: return best
        bd = any_limit
        for la, lo in all_pts:
            d = math.hypot((la - lat) * 111320, (lo - lon) * 111320 * math.cos(math.radians(lat)))
            if d < bd: best, bd = (la, lo), d
        return best

    added, pubs, new_keys = 0, set(), []
    for jid in all_jobs():
        dec = s3_get_json(f"{UPLOADS}{jid}/decisions.json", {}) or {}
        if not dec: continue
        props = s3_get_json(f"{UPLOADS}{jid}/proposals.json", []) or []
        st = s3_get_json(f"{UPLOADS}{jid}/status.json", {}) or {}
        pub = st.get("filename", jid)
        for i_str, d in dec.items():
            if d.get("verdict") != "accept": continue
            tag = f"{jid}#{i_str}"
            if tag in ledger: continue
            try: p = props[int(i_str)]
            except Exception: continue
            if p.get("lat") is None:
                log(f"    пропуск (нет координат): {p.get('locality')}"); continue
            lat, lon = round(float(p["lat"]), 5), round(float(p["lon"]), 5)   # как во всём корпусе
            near = snap(p.get("locality"), lat, lon)
            if near: lat, lon = near
            e0, e1 = elev(p.get("elevation"))
            src = f"{pub} (загружено пользователем)"
            ws.append([p.get("locality", ""), p.get("admin", "ND"), join(p.get("excavation")),
                       join(p.get("geomorph")), join(p.get("deposits")), join(p.get("raw_terms")),
                       thick(p, "studied"), thick(p, "visible"), thick(p, "borehole_depth"),
                       thick(p, "unspecified"), e0, e1, join(p.get("strat")), join(p.get("dating")),
                       join(p.get("source_kinds")), str(p.get("n_records", 1)), "1", src,
                       " || ".join(p.get("evidence", [])[:6]), lat, lon, "ok"])
            ledger[tag] = {"locality": p.get("locality"), "user": d.get("user", "anon")}
            new_keys.append(f"{round(float(lat),6)},{round(float(lon),6)}")
            pubs.add(pub); added += 1
            log(f"    + {p.get('locality')}  [{tag}]")
    if added:
        wb.save(xl)
        json.dump(ledger, open(lp, "w", encoding="utf-8"), ensure_ascii=False)
    return added, pubs, new_keys

# ---------- 3. досчёт только по новым координатам ----------
def new_coords(keys):
    out = []
    for k in keys:
        try:
            la, lo = k.split(","); out.append((round(float(la), 5), round(float(lo), 5)))
        except Exception: pass
    return list(dict.fromkeys(out))

def add_dem(coords):
    p = os.path.join(TMP, "dem_elevation.json")
    dem = json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    todo = [c for c in coords if f"{c[0]},{c[1]}" not in dem]
    if not todo: return 0
    url = "https://api.open-meteo.com/v1/elevation?" + urllib.parse.urlencode(
        {"latitude": ",".join(str(a) for a, _ in todo), "longitude": ",".join(str(b) for _, b in todo)})
    try:
        el = json.load(urllib.request.urlopen(url, timeout=40)).get("elevation", [])
        for (a, b), e in zip(todo, el): dem[f"{a},{b}"] = e
        json.dump(dem, open(p, "w", encoding="utf-8"), ensure_ascii=False)
        return len(el)
    except Exception as e:
        log("    DEM недоступен:", str(e)[:80]); return 0

RADIUS = {"isolated_dwelling": 800, "farm": 800, "locality": 1200, "hamlet": 1200,
          "village": 1500, "quarter": 1500, "neighbourhood": 1500, "suburb": 2500, "borough": 2500,
          "municipality": 3000, "town": 4000, "city": 8000, "administrative": 15000,
          "county": 25000, "district": 25000, "state": 40000, "region": 40000}

def add_radius(coords):
    p = os.path.join(TMP, "accuracy_radius.json")
    acc = json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    todo = [c for c in coords if f"{c[0]},{c[1]}" not in acc]
    n = 0
    for la, lo in todo:
        q = urllib.parse.urlencode({"lat": la, "lon": lo, "format": "json", "zoom": 14})
        r_ = 5000
        try:
            req = urllib.request.Request("https://nominatim.openstreetmap.org/reverse?" + q,
                                         headers={"User-Agent": UA})
            d = json.load(urllib.request.urlopen(req, timeout=30))
            r_ = RADIUS.get(d.get("addresstype") or d.get("type") or "", 5000)
        except Exception:
            pass
        acc[f"{la},{lo}"] = r_; n += 1
        time.sleep(1.05)
    if n: json.dump(acc, open(p, "w", encoding="utf-8"), ensure_ascii=False)
    return n

# ---------- 4. привязка страниц только для новых строк ----------
def norm(s):
    s = (s or "").lower().replace("ё", "е")
    return re.sub(r"[^0-9a-zа-я]+", " ", s).strip()

UPLOAD_MARK = re.compile(r"\s*\(загружено пользователем\)\s*$", re.I)

def find_page(pages, ev):
    ev = norm(ev)
    if len(ev) < 12: return None, None
    probe = ev[:60]
    for idx, txt in pages:
        if probe in txt: return idx, "exact"
    w = ev.split()
    for n in (8, 6, 4):
        fr = " ".join(w[:n])
        if len(fr) < 12: continue
        for idx, txt in pages:
            if fr in txt: return idx, "exact"
    tok = set(x for x in w if len(x) >= 5) or set(x for x in w if len(x) >= 4)
    if len(tok) < 3: return None, None
    best, sc0 = None, 0
    for idx, txt in pages:
        sc = len(tok & set(txt.split()))
        if sc > sc0: sc0, best = sc, idx
    return (best, "fuzzy") if sc0 >= max(3, int(0.6 * len(tok))) else (None, None)

def link_pages(new_keys):
    """Качает OCR только затронутых публикаций и дополняет page_links_v2.json."""
    import openpyxl
    if not new_keys: return 0
    ocr = {}      # имя публикации -> [(idx, нормализованный текст)]
    tok = None
    while True:
        kw = {"Bucket": BUCKET, "Prefix": "ocrtext/upload/", "MaxKeys": 1000}
        if tok: kw["ContinuationToken"] = tok
        r = s3.list_objects_v2(**kw)
        for o in r.get("Contents", []):
            d = s3_get_json(o["Key"], {}) or {}
            base = os.path.basename(str(d.get("path", "")).replace("\\", "/"))
            if base:
                ocr.setdefault(base.lower(), []).extend(
                    [(pg.get("idx", 0), norm(pg.get("text", ""))) for pg in d.get("pages", [])])
        tok = r.get("NextContinuationToken")
        if not r.get("IsTruncated"): break

    wb = openpyxl.load_workbook(os.path.join(TMP, "records_clean_geo_v2.xlsx"), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True); H = list(next(rows)); C = {n: i for i, n in enumerate(H)}
    want = set(new_keys)
    found = collections.defaultdict(lambda: collections.defaultdict(lambda: {"pages": set(), "approx": True}))
    for r in rows:
        la, lo = r[C["lat"]], r[C["lon"]]
        if la in (None, "", "ND"): continue
        try: key = f"{round(float(la),6)},{round(float(lo),6)}"
        except Exception: continue
        if key not in want: continue
        srcs = [s.strip() for s in str(r[C["sources"]] or "").split("|") if s.strip()]
        evs = [e.strip() for e in str(r[C["evidence"]] or "").split("||") if e.strip()]
        for src in srcs:
            pages = ocr.get(UPLOAD_MARK.sub("", src).lower())
            if not pages: continue
            for ev in evs:
                idx, meth = find_page(pages, ev)
                if idx is not None:
                    d = found[key][src]
                    d["pages"].add(idx + 1)
                    if meth == "exact": d["approx"] = False
    wb.close()

    p = os.path.join(TMP, "page_links_v2.json")
    pl = json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    added = 0
    for key, sm in found.items():
        cur = pl.get(key, []); have = {e["src"] for e in cur}
        for src, d in sm.items():
            if d["pages"] and src not in have:
                cur.append({"src": src, "pages": sorted(d["pages"]), "approx": d["approx"]}); added += 1
        if cur: pl[key] = cur
    if added: json.dump(pl, open(p, "w", encoding="utf-8"), ensure_ascii=False)
    return added

# ---------- 5. рендер сканов только новых публикаций ----------
def slug(name):
    import hashlib
    base = re.sub(r"[^0-9A-Za-zА-Яа-яЁё]+", "_", name.rsplit(".", 1)[0]).strip("_")[:52]
    return f"{base}_{hashlib.md5(name.encode('utf-8')).hexdigest()[:6]}"

def render_scans(pubs):
    if not pubs: return 0
    import fitz
    p = os.path.join(TMP, "scan_index.json")
    idx = json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    plp = os.path.join(TMP, "page_links_v2.json")
    pl = json.load(open(plp, encoding="utf-8")) if os.path.exists(plp) else {}
    need = collections.defaultdict(set)
    for lst in pl.values():
        for e in lst:
            src = UPLOAD_MARK.sub("", e["src"])
            if src in pubs: need[src].update(e["pages"])
    if not need: return 0

    src_key = {}
    for jid in all_jobs():
        st = s3_get_json(f"{UPLOADS}{jid}/status.json", {}) or {}
        fn = st.get("filename")
        if fn in need:
            ext = ("." + fn.rsplit(".", 1)[-1].lower()) if "." in fn else ".pdf"
            src_key[fn] = f"{UPLOADS}{jid}/source{ext}"

    n = 0
    for src, pages in need.items():
        key = src_key.get(src)
        if not key: continue
        try:
            blob = s3.get_object(Bucket=BUCKET, Key=key)["Body"].read()
            doc = fitz.open(stream=blob, filetype="pdf")
        except Exception as e:
            log("    скан не отрендерен:", str(e)[:80]); continue
        sl = slug(src); got = idx.get(src, {})
        for pg in sorted(pages):
            i = pg - 1
            if i < 0 or i >= doc.page_count or str(pg) in got: continue
            k = f"scans/{sl}/p{pg}.jpg"
            s3.put_object(Bucket=SITE_BUCKET, Key=k,
                          Body=doc[i].get_pixmap(dpi=120).tobytes("jpeg", jpg_quality=72),
                          ContentType="image/jpeg", CacheControl="public, max-age=2592000")
            got[str(pg)] = k; n += 1
        doc.close(); idx[src] = got
    if n: json.dump(idx, open(p, "w", encoding="utf-8"), ensure_ascii=False)
    return n

# ---------- 6. сборка и публикация карты ----------
def build_map():
    os.environ["DATA_DIR"] = TMP
    os.environ["OUT_DIR"] = "/tmp"
    os.environ["TEMPLATE"] = os.path.join(TMP, "template.html")
    os.environ["SRC_XLSX"] = os.path.join(TMP, "records_clean_geo_v2.xlsx")
    # сборщик берём из пакета, а если его туда не положили — из бакета (data/builder.py).
    # Второй путь удобнее: правку сборщика достаточно залить через 62_sync_data.py, без передеплоя функции.
    here = os.path.dirname(os.path.abspath(__file__))
    b = os.path.join(here, "builder.py")
    if not os.path.exists(b):
        b = "/tmp/builder.py"
        open(b, "wb").write(s3.get_object(Bucket=BUCKET, Key=DATA_PREFIX + "builder.py")["Body"].read())
        log("    сборщик взят из бакета")
    runpy.run_path(b, run_name="__main__")
    return "/tmp/index.html"

def markers_in(path):
    try:
        h = open(path, encoding="utf-8").read()
        m = re.search(r"const DATA = (\[.*?\]);\n", h, re.S)
        return len(json.loads(m.group(1))) if m else None
    except Exception:
        return None

CORS = {"Access-Control-Allow-Origin": "*", "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type"}

def _resp(code, obj):
    return {"statusCode": code, "headers": {**CORS, "Content-Type": "application/json"},
            "body": json.dumps(obj, ensure_ascii=False)}

def handler(event, context):
    # кнопку жмёт браузер, поэтому нужен предварительный запрос OPTIONS и заголовки CORS
    if str((event or {}).get("httpMethod") or "POST").upper() == "OPTIONS":
        return _resp(200, {"ok": True})
    body = {}
    raw = (event or {}).get("body")
    if raw:
        try: body = json.loads(base64.b64decode(raw).decode() if event.get("isBase64Encoded") else raw)
        except Exception: body = {}
    force, dry = bool(body.get("force")), bool(body.get("dry"))
    t0 = time.time()

    pull_data()
    log("[2] сливаю принятые объекты...")
    added, pubs, new_keys = merge_accepted()
    log(f"    добавлено: {added}")
    if not added and not force:
        return _resp(200, {"added": 0, "msg": "новых принятых объектов нет"})

    coords = new_coords(new_keys)
    log(f"[3] досчёт по новым координатам: {len(coords)}")
    log(f"    высот: {add_dem(coords)} | радиусов: {add_radius(coords)}")
    log(f"[4] привязка страниц: +{link_pages(new_keys)}")
    log(f"[5] сканы: +{render_scans(pubs)} страниц")

    before = markers_in("/tmp/index.html")
    log("[6] собираю карту...")
    out = build_map()
    after = markers_in(out)
    log(f"    маркеров: {after}")

    published = False
    if dry:
        log("    сухой прогон — не публикую")
    elif before and after and after < before * MIN_KEEP:
        log(f"    СТОП: маркеров стало меньше {MIN_KEEP:.0%} от прежнего, не публикую")
    else:
        s3.put_object(Bucket=SITE_BUCKET, Key="index.html", Body=open(out, "rb").read(),
                      ContentType="text/html; charset=utf-8")
        push_data(["records_clean_geo_v2.xlsx", "dem_elevation.json", "accuracy_radius.json",
                   "page_links_v2.json", "scan_index.json", "merged_uploads.json"])
        published = True
        log("    опубликовано")

    return _resp(200, {"added": added, "markers": after, "published": published,
                       "sec": round(time.time() - t0, 1)})
