# Вносим площадки и выработки, размещённые скриптом 72, в мастер-таблицу — чтобы они появились
# на карте отдельными точками, как просил геолог (п.3): «конкретные выработки как отдельные точки,
# сводный разрез как обобщающий объект, связанный с ними».
#
# Связь хранится отдельно (object_links.json), чтобы карточка могла показать оба направления:
# у сводного разреза — список входящих площадок, у площадки — ссылку на сводный.
#
#   DRY=1 python 73_merge_sublevels.py
import os, re, sys, json, shutil, pathlib, collections
import openpyxl, unicodedata
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "records_clean_geo_v2.xlsx"
SUB = HERE / "sublevel_objects.json"
LINKS = HERE / "object_links.json"
DRY = os.environ.get("DRY", "0") == "1"

sub = json.load(open(SUB, encoding="utf-8"))
placed = [v for v in sub.values() if v.get("lat") is not None]
print(f"размещённых объектов: {len(placed)} из {len(sub)}")

wb = openpyxl.load_workbook(XLSX); ws = wb.active
H = [c.value for c in ws[1]]; C = {n: i for i, n in enumerate(H)}

# Уже существующие точки: если объект уже есть на этой координате под тем же именем — не дублируем
have = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    la, lo = row[C["lat"]], row[C["lon"]]
    if str(la) in ("None", "", "ND"): continue
    have.add((str(row[C["nearest_locality"]] or "").strip().lower(),
              round(float(la), 5), round(float(lo), 5)))

# «borehole_depth: 1.8-5.5 м» -> в нужную колонку мощности
TH_MAP = {"borehole_depth": "thickness_borehole_depth", "visible": "thickness_visible",
          "studied": "thickness_studied", "unspecified": "thickness_unspecified"}


def thickness(vals):
    out = collections.defaultdict(list)
    for v in vals or []:
        m = re.match(r"\s*(\w+)\s*:\s*(.+?)\s*м?\s*$", str(v))
        if m and m.group(1) in TH_MAP: out[TH_MAP[m.group(1)]].append(m.group(2))
    return {k: "; ".join(dict.fromkeys(v)) for k, v in out.items()}


def plain(s):
    """Без знака ударения: в извлечённых названиях встречается «Отка́зное» из оригинала, и по нему
    не сходится ни поиск родителя, ни показ в карточке."""
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "").strip())
                   if not unicodedata.combining(c))


links = json.load(open(LINKS, encoding="utf-8")) if LINKS.exists() else {}
added, skipped = 0, 0
import math
# Уже занятые координаты: новая точка ближе 300 м к существующей — это не новый объект, а второе
# имя того же места («Лихвинский разрез» = «Чекалин») или выработка внутри площадки. Такие на карту
# отдельными маркерами не выносим: они всё равно слиплись бы в один и только зашумили названия.
occupied = [(la, lo) for _, la, lo in have]
MIN_APART = float(os.environ.get("MIN_APART", 300))


def too_close(la, lo):
    for a, b in occupied:
        if math.hypot((la - a) * 111320, (lo - b) * 111320 * math.cos(math.radians(la))) < MIN_APART:
            return True
    return False


for v in placed:
    key = (v["name"].strip().lower(), round(v["lat"], 5), round(v["lon"], 5))
    if key in have or v.get("relation") == "то же самое место" or too_close(v["lat"], v["lon"]):
        skipped += 1
        print(f"   пропуск ({v.get('relation')}): {v['name']} — точка не новая")
        continue
    occupied.append((v["lat"], v["lon"]))
    row = [None] * len(H)
    row[C["nearest_locality"]] = v["name"]
    row[C["evidence"]] = " || ".join(v.get("evidence") or [])
    row[C["lat"]], row[C["lon"]] = v["lat"], v["lon"]
    # координата вычислена от соседней точки по описанию в тексте, а не найдена напрямую
    row[C["loc_confidence"]] = "uncertain"
    row[C["excavation_type"]] = "; ".join(v.get("excavations") or []) or None
    row[C["n_records"]] = 1
    row[C["n_sources"]] = 1
    row[C["source_kinds"]] = "prose"
    for k, val in thickness(v.get("thickness")).items(): row[C[k]] = val
    ws.append(row)
    have.add(key); added += 1
    print(f"   + {v['name'][:22]:22} {v['lat']},{v['lon']}  выработки: {', '.join(v.get('excavations') or []) or '—'}")

    # двусторонняя связь: у площадки — родитель, у родителя — список входящих
    mk = f"{v['lat']},{v['lon']}"
    par = v.get("inherited_from") or v.get("parent") or ""
    links.setdefault(mk, {})["partOf"] = par
    links[mk]["excavations"] = v.get("excavations") or []
    links[mk]["why"] = v.get("why", "")
    for p in placed + []:
        pass
# обратная связь: родитель -> входящие площадки (по имени, координату родителя ищем в таблице)
byname, pretty = {}, {}
for row in ws.iter_rows(min_row=2, values_only=True):
    la, lo = row[C["lat"]], row[C["lon"]]
    if str(la) in ("None", "", "ND"): continue
    nm = plain(row[C["nearest_locality"]])
    byname.setdefault(nm.lower(), f"{round(float(la),5)},{round(float(lo),5)}")
    pretty.setdefault(nm.lower(), nm)
# Родитель для списка «включает» — не тот, от кого считали координату, а настоящий объект.
# Site 2 позиционно отсчитан от Site 1, но обе площадки принадлежат сводному разрезу «Отказное»,
# поэтому поднимаемся по цепочке, пока не выйдем за пределы новых объектов.
# Звенья цепи — только НОВЫЕ площадки. Объект с отношением «то же самое место» это просто второе
# имя существующего маркера; если считать его звеном, подъём упирается в самоссылку и связь теряется.
sub_names = {plain(v["name"]).lower() for v in placed
             if v.get("relation") != "то же самое место"}
def key_added(v):
    """Ссылку даём только на объект, который реально есть в таблице отдельной строкой, — иначе
    клик ведёт в никуда. Проверяем по факту, а не по тому, добавили ли мы его в этом прогоне:
    при повторном запуске он уже существует."""
    return (v["name"].strip().lower(), round(v["lat"], 5), round(v["lon"], 5)) in have


def real_parent(v, depth=0):
    par = plain(v.get("parent") or v.get("inherited_from")).lower()
    if not par or par == plain(v["name"]).lower() or depth > 4:
        return ""
    if par in sub_names:                       # родитель — тоже новая площадка, идём выше
        nxt = next((x for x in placed if plain(x["name"]).lower() == par), None)
        return real_parent(nxt, depth + 1) if nxt else par
    return par


for v in placed:
    par = real_parent(v)
    pk = byname.get(par)
    if not pk or pk == f"{round(v['lat'],5)},{round(v['lon'],5)}": continue   # не ссылаемся на себя
    e = links.setdefault(pk, {})
    e.setdefault("includes", [])
    item = {"name": v["name"], "at": f"{v['lat']},{v['lon']}"}
    if item not in e["includes"] and plain(v["name"]).lower() != par and key_added(v):
        e["includes"].append(item)
    links.setdefault(f"{v['lat']},{v['lon']}", {})["partOf"] = pretty.get(par, par)

print(f"\nдобавлено строк: {added} | пропущено как дубли: {skipped}")
print(f"связей записано: {len(links)}")
if DRY:
    print("сухой прогон — ничего не сохранено"); sys.exit(0)
shutil.copy(XLSX, XLSX.with_suffix(".before_sublevels.xlsx"))
wb.save(XLSX)
json.dump(links, open(LINKS, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"-> {XLSX.name} и {LINKS.name} обновлены (бэкап .before_sublevels.xlsx)")
