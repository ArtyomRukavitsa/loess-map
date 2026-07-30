# Уточнение координат по КОНТЕКСТУ из самих цитат (замечание геолога, п.2 третьей итерации).
# Проблема: приоритет отдавался названию села, а одноимённых сёл много. При этом регион нередко
# написан прямо в цитате — например «Тамбовская область, Сосновский район, с. Березовка, р. Грязнушка»,
# а точка при этом стояла в другом конце страны.
# Здесь: достаём из текста область/край/республику и район, берём их границы через Nominatim
# (кэшируется), и если точка вне этих границ — ищем населённый пункт заново ВНУТРИ них.
# Диагностика:  DRY=1 python 64_fix_geocode_context.py
# Правка:       python 64_fix_geocode_context.py
import openpyxl, json, time, sys, os, re, math, shutil, pathlib, urllib.request, urllib.parse
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "records_clean_geo_v2.xlsx"
BBOX_CACHE = HERE / "geocache_bbox.json"
SEC_CACHE = HERE / "geocache_sections.json"
DRY = os.environ.get("DRY", "0") == "1"
UA = "loess-sections/1.0 (research)"
CIS = "ru,ua,kz,by,md,ge,az,am,uz,kg,tj,tm,ee,lv,lt"
SETTLE = {"city", "town", "village", "hamlet", "municipality", "administrative", "suburb",
          "borough", "isolated_dwelling", "locality", "quarter"}

# --- вытаскиваем административный контекст из текста ---
# «Тамбовской области» / «Тамбовская обл.» -> «Тамбовская область»; то же для края и республики
RE_OBL = re.compile(r"([А-ЯЁ][а-яё\-]+ск)(?:ая|ой|ую|ом)\s+(?:област[ьия]|обл\.)")
RE_KRAI = re.compile(r"([А-ЯЁ][а-яё\-]+ск)(?:ий|ого|ом)\s+кра[йея]")
RE_RESP = re.compile(r"[Рр]еспублик[аиеу]\s+([А-ЯЁ][а-яё\-]+)")
RE_RAION = re.compile(r"([А-ЯЁ][а-яё\-]+ск)(?:ий|ого|ом)\s+район")

def admin_from_text(t):
    """Возвращает (регион, район) в именительном падеже, если нашлись в тексте."""
    t = str(t or "")
    reg = None
    m = RE_OBL.search(t)
    if m: reg = m.group(1) + "ая область"
    if not reg:
        m = RE_KRAI.search(t)
        if m: reg = m.group(1) + "ий край"
    if not reg:
        m = RE_RESP.search(t)
        if m: reg = "Республика " + m.group(1)
    m = RE_RAION.search(t)
    rai = (m.group(1) + "ий район") if m else None
    return reg, rai

def req(url):
    r = urllib.request.Request(url, headers={"User-Agent": UA})
    return json.load(urllib.request.urlopen(r, timeout=30))

bbox_cache = json.load(open(BBOX_CACHE, encoding="utf-8")) if BBOX_CACHE.exists() else {}
sec_cache = json.load(open(SEC_CACHE, encoding="utf-8")) if SEC_CACHE.exists() else {}
_req_n = [0]

def region_bbox(name):
    """Границы региона по названию (кэшируется — регионов десятки, а строк тысячи)."""
    if name in bbox_cache: return bbox_cache[name]
    res = None
    try:
        for it in req("https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
                {"q": name, "format": "json", "limit": 3, "countrycodes": CIS})):
            bb = it.get("boundingbox")
            if bb and it.get("class") in ("boundary", "place"):
                res = [float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3])]  # lat0,lat1,lon0,lon1
                break
    except Exception:
        res = None
    bbox_cache[name] = res; _req_n[0] += 1; time.sleep(1.05); return res

def geocode_in(name, bbox):
    """Ищем населённый пункт строго внутри области/района."""
    key = f"{name}|bbox|{bbox}"
    if key in sec_cache: return sec_cache[key]
    la0, la1, lo0, lo1 = bbox; res = None
    try:
        for it in req("https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
                {"q": name, "format": "json", "limit": 8, "countrycodes": CIS,
                 "viewbox": f"{lo0},{la1},{lo1},{la0}", "bounded": 1})):
            if it.get("class") in ("place", "boundary") and it.get("type") in SETTLE:
                la, lo = float(it["lat"]), float(it["lon"])
                if la0 <= la <= la1 and lo0 <= lo <= lo1: res = [la, lo]; break
    except Exception:
        res = None
    sec_cache[key] = res; _req_n[0] += 1; time.sleep(1.05); return res

def inbox(la, lo, b, pad=0.15):
    return (b[0] - pad) <= la <= (b[1] + pad) and (b[2] - pad) <= lo <= (b[3] + pad)

# --- проход по таблице ---
wb = openpyxl.load_workbook(XLSX); ws = wb.active
H = [c.value for c in ws[1]]; C = {n: H.index(n) for n in H}
data = [list(r) for r in list(ws.values)[1:]]

cand = []
for i, r in enumerate(data):
    la, lo = r[C["lat"]], r[C["lon"]]
    if la in (None, "", "ND") or str(la) == "None": continue
    ctx = " ".join(str(r[C[k]] or "") for k in ("evidence", "geomorphic_position", "administrative_unit"))
    reg, rai = admin_from_text(ctx)
    if reg: cand.append((i, reg, rai))

print(f"строк с координатами: {sum(1 for r in data if str(r[C['lat']]) not in ('None','','ND'))}")
print(f"из них регион назван прямо в тексте: {len(cand)}")

if DRY:
    from collections import Counter
    print("\nтоп упомянутых регионов:")
    for reg, n in Counter(x[1] for x in cand).most_common(12):
        print(f"   {n:4}  {reg}")
    print("\nпримеры:")
    for i, reg, rai in cand[:12]:
        r = data[i]
        print(f"   {str(r[C['nearest_locality']])[:24]:24} ({r[C['lat']]}, {r[C['lon']]}) <- {reg}" +
              (f", {rai}" if rai else ""))
    sys.exit(0)

n_ok = n_fix = n_drop = n_nobbox = 0
for k, (i, reg, rai) in enumerate(cand, 1):
    r = data[i]
    bb = region_bbox(reg)
    if not bb: n_nobbox += 1; continue
    la, lo = float(r[C["lat"]]), float(r[C["lon"]])
    if inbox(la, lo, bb): n_ok += 1; continue           # точка и так в названном регионе
    name = str(r[C["nearest_locality"]]).strip()
    # сначала пробуем внутри района (точнее), потом внутри области
    new = None
    if rai:
        bbr = region_bbox(f"{rai}, {reg}")
        if bbr: new = geocode_in(name, bbr)
    if not new: new = geocode_in(name, bb)
    if new:
        r[C["lat"]], r[C["lon"]] = round(new[0], 5), round(new[1], 5)
        if not str(r[C["administrative_unit"]] or "").strip() or r[C["administrative_unit"]] == "ND":
            r[C["administrative_unit"]] = reg           # заодно проставляем найденный регион
        n_fix += 1
        if n_fix <= 25:
            print(f"  ИСПРАВЛЕНО {name}: {la:.3f},{lo:.3f} -> {new[0]:.3f},{new[1]:.3f}  ({reg})")
    else:
        r[C["lat"]], r[C["lon"]] = "ND", "ND"           # текст противоречит координате, замены нет
        r[C["loc_confidence"]] = "uncertain"
        n_drop += 1
    if k % 25 == 0:
        json.dump(bbox_cache, open(BBOX_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
        json.dump(sec_cache, open(SEC_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
        print(f"  ...{k}/{len(cand)} (исправлено {n_fix}, снято {n_drop}, запросов {_req_n[0]})", flush=True)

json.dump(bbox_cache, open(BBOX_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
json.dump(sec_cache, open(SEC_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
shutil.copy(XLSX, XLSX.with_suffix(".before_context.xlsx"))
wb2 = openpyxl.Workbook(); w = wb2.active; w.title = "sections"; w.append(H)
for r in data: w.append(r)
wb2.save(XLSX)
print(f"\nИТОГ: в своём регионе {n_ok} | исправлено {n_fix} | снято {n_drop} | "
      f"без границ региона {n_nobbox} | запросов {_req_n[0]}")
print(f"-> {XLSX.name} обновлён (бэкап .before_context.xlsx)")
