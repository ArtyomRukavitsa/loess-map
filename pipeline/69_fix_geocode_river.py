# Уточнение координат по РЕКЕ из цитат (просьба геолога: «стоит сильнее учитывать бассейн и название
# реки»; в правках проверяющих это самая частая подсказка — «на Десне», «в долине реки Еи», «на берегу Оки»).
#
# Логика та же, что в 64 для областей: достаём ориентир из текста, берём его границы через Nominatim
# и, если точка вне них, ищем населённый пункт заново ВНУТРИ. Отличие — река тянется на сотни
# километров, поэтому её рамка широкая: она отсекает грубые промахи (Мезин в Саратовской вместо
# Черниговской), но не уточняет位置 внутри бассейна. Это ровно та ошибка, на которую жалуются.
#
# Диагностика:  DRY=1 python 69_fix_geocode_river.py
import openpyxl, json, time, sys, os, re, math, shutil, pathlib, urllib.request, urllib.parse, collections
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "records_clean_geo_v2.xlsx"
CACHE = HERE / "geocache_river.json"
SEC_CACHE = HERE / "geocache_sections.json"
DRY = os.environ.get("DRY", "0") == "1"
UA = "loess-sections/1.0 (research)"
CIS = "ru,ua,kz,by,md,ge,az,am,uz,kg,tj,tm,ee,lv,lt"
SETTLE = {"city", "town", "village", "hamlet", "municipality", "administrative", "suburb",
          "borough", "isolated_dwelling", "locality", "quarter"}

# «р. Десна», «реки Еи», «на Десне», «долина р. Оки», «бассейн реки Дон»
RIVER_RE = re.compile(
    r"(?:р\.|реки|реке|рекой|река|бассейн\w*\s+р(?:еки|\.)|долин\w*\s+р(?:еки|\.)|берег\w*\s+р(?:еки|\.))"
    r"\s*([А-ЯЁ][а-яё\-]{2,})", re.U)
# «на Десне», «на Оке» — без слова «река»
RIVER_ON = re.compile(r"\bна\s+([А-ЯЁ][а-яё\-]{3,}[еиы])\b", re.U)

def variants(w):
    """Приводим склонённое название к именительному: «Десне»→«Десна», «Оки»→«Ока», «Еи»→«Ея»."""
    w = w.strip()
    out = [w]
    if len(w) > 2:                     # «Оки»→«Ока»: короткие названия тоже склоняются
        stem = w[:-1]
        for suf in ("а", "я"):
            out.append(stem + suf)
    if w.endswith("ой") or w.endswith("ей"):
        out.append(w[:-2] + "а")
    seen, res = set(), []
    for v in out:
        if v.lower() not in seen: seen.add(v.lower()); res.append(v)
    return res[:3]

cache = json.load(open(CACHE, encoding="utf-8")) if CACHE.exists() else {}
sec_cache = json.load(open(SEC_CACHE, encoding="utf-8")) if SEC_CACHE.exists() else {}
_req = [0]

def req(url):
    r = urllib.request.Request(url, headers={"User-Agent": UA})
    return json.load(urllib.request.urlopen(r, timeout=30))

def river_bbox(name):
    """Границы реки. Берём только водные объекты, иначе «Ея» найдётся как село."""
    if name in cache: return cache[name]
    res = None
    for v in variants(name):
        try:
            for it in req("https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
                    {"q": v + " река", "format": "json", "limit": 5, "countrycodes": CIS})):
                cls, typ = it.get("class"), it.get("type")
                if cls == "waterway" or (cls == "natural" and typ in ("water", "stream")):
                    bb = it.get("boundingbox")
                    if bb:
                        res = [float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3])]
                        break
        except Exception:
            res = None
        _req[0] += 1; time.sleep(1.05)
        if res: break
    cache[name] = res
    return res

def geocode_in(name, bbox):
    key = f"{name}|riv|{bbox}"
    if key in sec_cache: return sec_cache[key]
    la0, la1, lo0, lo1 = bbox; res = None
    try:
        for it in req("https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
                {"q": name, "format": "json", "limit": 8, "countrycodes": CIS,
                 "viewbox": f"{lo0},{la1},{lo1},{la0}", "bounded": 1})):
            if it.get("class") in ("place", "boundary") and it.get("type") in SETTLE:
                la, lo = float(it["lat"]), float(it["lon"])
                if la0 <= la <= la1 and lo0 <= lo <= lo1: res = [la, lo]; break
    except Exception:
        res = None
    sec_cache[key] = res; _req[0] += 1; time.sleep(1.05); return res

def inbox(la, lo, b, pad=0.3):
    return (b[0] - pad) <= la <= (b[1] + pad) and (b[2] - pad) <= lo <= (b[3] + pad)

wb = openpyxl.load_workbook(XLSX); ws = wb.active
H = [c.value for c in ws[1]]; C = {n: H.index(n) for n in H}
data = [list(r) for r in list(ws.values)[1:]]

cand = []
for i, r in enumerate(data):
    la = r[C["lat"]]
    if la in (None, "", "ND") or str(la) == "None": continue
    ctx = " ".join(str(r[C[k]] or "") for k in ("evidence", "geomorphic_position", "administrative_unit"))
    m = RIVER_RE.search(ctx) or RIVER_ON.search(ctx)
    if m: cand.append((i, m.group(1)))

print(f"строк с координатами: {sum(1 for r in data if str(r[C['lat']]) not in ('None','','ND'))}")
print(f"из них река названа в тексте: {len(cand)}")
top = collections.Counter(x[1] for x in cand)
print("чаще всего:", ", ".join(f"{n}×{c}" for n, c in top.most_common(10)))

if DRY:
    print("\nпримеры:")
    for i, riv in cand[:12]:
        r = data[i]
        print(f"   {str(r[C['nearest_locality']])[:22]:22} ({r[C['lat']]}, {r[C['lon']]}) <- река {riv} -> {variants(riv)}")
    sys.exit(0)

MAX_SPAN = float(os.environ.get("MAX_SPAN", 6.0))   # градусов; шире — река не ограничивает
n_ok = n_fix = n_flag = n_nobb = n_wide = 0
for k, (i, riv) in enumerate(cand, 1):
    r = data[i]
    bb = river_bbox(riv)
    if not bb: n_nobb += 1; continue
    # Крупные реки как ограничение бесполезны: рамка Лены 19°x25°, Волги 12°x18° — поиск внутри
    # почти не ограничен и переносит точку к случайному тёзке за тысячу километров (проверено
    # на «Ивановке»: Днепр увёл её под Одессу). Для таких рек только помечаем, не двигаем.
    if max(bb[1] - bb[0], bb[3] - bb[2]) > MAX_SPAN:
        la_, lo_ = float(r[C["lat"]]), float(r[C["lon"]])
        if not inbox(la_, lo_, bb):
            r[C["loc_confidence"]] = "uncertain"; n_flag += 1; n_wide += 1
        else:
            n_ok += 1
        continue
    la, lo = float(r[C["lat"]]), float(r[C["lon"]])
    if inbox(la, lo, bb): n_ok += 1; continue
    name = str(r[C["nearest_locality"]]).strip()
    new = geocode_in(name, bb)
    if new and math.hypot((new[0] - la) * 111320, (new[1] - lo) * 111320 * math.cos(math.radians(la))) > 50000:
        r[C["lat"]], r[C["lon"]] = round(new[0], 5), round(new[1], 5)
        n_fix += 1
        if n_fix <= 25: print(f"  ИСПРАВЛЕНО {name}: {la:.3f},{lo:.3f} -> {new[0]:.3f},{new[1]:.3f}  (река {riv})")
    else:
        r[C["loc_confidence"]] = "uncertain"; n_flag += 1
    if k % 25 == 0:
        json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)
        json.dump(sec_cache, open(SEC_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
        print(f"  ...{k}/{len(cand)} (исправлено {n_fix}, помечено {n_flag}, запросов {_req[0]})", flush=True)

json.dump(cache, open(CACHE, "w", encoding="utf-8"), ensure_ascii=False)
json.dump(sec_cache, open(SEC_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
shutil.copy(XLSX, XLSX.with_suffix(".before_river.xlsx"))
wb2 = openpyxl.Workbook(); w = wb2.active; w.title = "sections"; w.append(H)
for r in data: w.append(r)
wb2.save(XLSX)
print(f"\nИТОГ: в бассейне {n_ok} | исправлено {n_fix} | помечено сомнительными {n_flag} | "
      f"река не найдена {n_nobb} | запросов {_req[0]}")
