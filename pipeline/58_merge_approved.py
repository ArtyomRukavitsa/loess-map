# Перенос принятых при проверке объектов из загруженных публикаций в основную таблицу карты.
# Читает uploads/<job>/proposals.json + decisions.json из Object Storage, берёт только verdict=accept,
# дописывает строки в records_clean_geo_v2.xlsx (та же схема колонок -> карта собирается без изменений).
# Повторные запуски безопасны: перенесённые помечаются в merged_uploads.json и второй раз не добавляются.
# Запуск:  python 58_merge_approved.py         (сухой прогон: DRY=1 python 58_merge_approved.py)
import os, sys, json, re, shutil, pathlib
import openpyxl, boto3
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
DRY = os.environ.get("DRY", "0") == "1"
XLSX = HERE / "records_clean_geo_v2.xlsx"
LEDGER = HERE / "merged_uploads.json"

sec = {}
for line in open(HERE / ".secrets", encoding="utf-8"):
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1); sec[k.strip()] = v.strip()
s3 = boto3.client("s3", endpoint_url="https://storage.yandexcloud.net", region_name="ru-central1",
                  aws_access_key_id=sec["AWS_ACCESS_KEY_ID"], aws_secret_access_key=sec["AWS_SECRET_ACCESS_KEY"])
BUCKET = sec.get("BUCKET", "loess-results")
PREFIX = "uploads/"

def s3_get(key, default=None):
    try: return json.loads(s3.get_object(Bucket=BUCKET, Key=key)["Body"].read())
    except Exception: return default

def join(v):
    if isinstance(v, list): return "; ".join(str(x) for x in v if x not in (None, "", "ND"))
    return "" if v in (None, "ND") else str(v)

def thick(props, kind):                       # "studied: 12 м" -> "12" для нужного типа
    out = []
    for t in props.get("thickness", []):
        m = re.match(r"\s*(\w+)\s*:\s*(.+?)\s*м?\s*$", str(t))
        if m and m.group(1) == kind: out.append(m.group(2))
    return "; ".join(out)

def elev(v):
    m = re.findall(r"-?\d+(?:[.,]\d+)?", str(v or ""))
    return (m[0].replace(",", "."), m[-1].replace(",", ".")) if m else ("", "")

ledger = json.load(open(LEDGER, encoding="utf-8")) if LEDGER.exists() else {}

# --- индекс уже имеющихся точек: новый объект того же места «прилипает» к существующей координате,
#     иначе на карте появляется дубль-маркер в нескольких метрах от старого ---
import math, collections
_wb0 = openpyxl.load_workbook(XLSX, read_only=True); _ws0 = _wb0.active
_rows0 = _ws0.iter_rows(values_only=True); _h0 = list(next(_rows0)); _c0 = {n: i for i, n in enumerate(_h0)}
existing = collections.defaultdict(list)
for _r in _rows0:
    _la, _lo = _r[_c0["lat"]], _r[_c0["lon"]]
    if _la in (None, "", "ND") or str(_la) == "None": continue
    try: existing[str(_r[_c0["nearest_locality"]] or "").strip().lower()].append((float(_la), float(_lo)))
    except Exception: pass
_wb0.close()

_all_pts = [q for pts in existing.values() for q in pts]

def snap(loc, lat, lon, limit_m=300, any_m=150):
    """Совмещаем с уже имеющейся точкой: сначала по совпадению названия, затем — по близости
    с любой точкой. Второе нужно из-за разного написания («Отказное» и «Otkaznoe» — одно место)."""
    best, bestd = None, limit_m
    for la, lo in existing.get(str(loc or "").strip().lower(), []):
        d = math.hypot((la - lat) * 111320, (lo - lon) * 111320 * math.cos(math.radians(lat)))
        if d < bestd: best, bestd = (la, lo), d
    if best: return best
    bestd = any_m
    for la, lo in _all_pts:
        d = math.hypot((la - lat) * 111320, (lo - lon) * 111320 * math.cos(math.radians(lat)))
        if d < bestd: best, bestd = (la, lo), d
    return best

# --- собрать принятые объекты по всем задачам ---
jobs, tok = [], None
while True:
    kw = {"Bucket": BUCKET, "Prefix": PREFIX, "MaxKeys": 1000}
    if tok: kw["ContinuationToken"] = tok
    r = s3.list_objects_v2(**kw)
    jobs += [o["Key"].split("/")[1] for o in r.get("Contents", []) if o["Key"].endswith("/status.json")]
    tok = r.get("NextContinuationToken")
    if not r.get("IsTruncated"): break

rows, seen_new = [], []
for jid in sorted(set(jobs)):
    dec = s3_get(f"{PREFIX}{jid}/decisions.json", {}) or {}
    if not dec: continue
    props = s3_get(f"{PREFIX}{jid}/proposals.json", []) or []
    st = s3_get(f"{PREFIX}{jid}/status.json", {}) or {}
    pub = st.get("filename", jid)
    for i_str, d in dec.items():
        if d.get("verdict") != "accept": continue
        tag = f"{jid}#{i_str}"
        if tag in ledger: continue
        try: p = props[int(i_str)]
        except Exception: continue
        if p.get("lat") is None:
            print(f"  пропуск (нет координат): {p.get('locality')} [{tag}]"); continue
        lat, lon = round(float(p["lat"]), 5), round(float(p["lon"]), 5)   # как во всём корпусе
        near = snap(p.get("locality"), lat, lon)
        if near:
            lat, lon = near                       # слить с существующей точкой того же места
            print(f"  ~ {p.get('locality')}: координата совмещена с уже имеющейся точкой")
        e_min, e_max = elev(p.get("elevation"))
        rows.append([
            p.get("locality", ""), p.get("admin", "ND"), join(p.get("excavation")), join(p.get("geomorph")),
            join(p.get("deposits")), join(p.get("raw_terms")),
            thick(p, "studied"), thick(p, "visible"), thick(p, "borehole_depth"), thick(p, "unspecified"),
            e_min, e_max, join(p.get("strat")), join(p.get("dating")), join(p.get("source_kinds")),
            str(p.get("n_records", 1)), "1", f"{pub} (загружено пользователем)",
            " || ".join(p.get("evidence", [])[:6]), lat, lon, "ok",
        ])
        seen_new.append((tag, p.get("locality"), d.get("user", "anon")))

if not rows:
    print("новых принятых объектов нет"); sys.exit(0)

print(f"принято к переносу: {len(rows)}")
for tag, loc, who in seen_new: print(f"  + {loc}  [{tag}, проверил: {who}]")
if DRY:
    print("сухой прогон — таблица не изменена"); sys.exit(0)

shutil.copy(XLSX, XLSX.with_suffix(".before_uploads.xlsx"))     # бэкап перед записью
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
hdr = [c.value for c in ws[1]]
if len(hdr) != 22:
    print(f"ВНИМАНИЕ: ожидалось 22 колонки, в файле {len(hdr)} — проверь схему"); sys.exit(1)
for r in rows: ws.append(r)
wb.save(XLSX)
for tag, loc, who in seen_new: ledger[tag] = {"locality": loc, "user": who}
json.dump(ledger, open(LEDGER, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"дописано строк: {len(rows)} -> {XLSX.name} (бэкап .before_uploads.xlsx)")
print("дальше: пересобрать карту (build_customer_map_v2.py)")
