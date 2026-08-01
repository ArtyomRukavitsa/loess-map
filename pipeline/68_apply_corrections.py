# Применение правок проверяющих: если в комментарии указаны правильные координаты — переносим точку.
# Замечание коллег приходит в виде «Правильные координаты. 43.316635, 43.605317» — такие правки
# можно применять машинально, они однозначны. Остальные (без координат) остаются для разбора руками.
#
# Повторные запуски безопасны: применённые отмечаются в applied_corrections.json.
#   DRY=1 python 68_apply_corrections.py   — только показать, что изменится
import os, re, sys, json, math, shutil, pathlib, collections
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "records_clean_geo_v2.xlsx"
LEDGER = HERE / "applied_corrections.json"
DRY = os.environ.get("DRY", "0") == "1"

sec = {}
for line in open(HERE / ".secrets", encoding="utf-8"):
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, v = line.split("=", 1); sec[k.strip()] = v.strip()
import boto3
s3 = boto3.client("s3", endpoint_url="https://storage.yandexcloud.net", region_name="ru-central1",
                  aws_access_key_id=sec["AWS_ACCESS_KEY_ID"], aws_secret_access_key=sec["AWS_SECRET_ACCESS_KEY"])
BUCKET = sec.get("BUCKET", "loess-results")

# «43.316635, 43.605317» — пара десятичных чисел в правдоподобных пределах
COORD = re.compile(r"(-?\d{2}\.\d{3,})\s*[,;]\s*(-?\d{1,3}\.\d{3,})")

def valid(la, lo):
    return 35 <= la <= 82 and 19 <= lo <= 180

# ---------- читаем правки ----------
keys, tok = [], None
while True:
    kw = {"Bucket": BUCKET, "Prefix": "corrections/", "MaxKeys": 1000}
    if tok: kw["ContinuationToken"] = tok
    r = s3.list_objects_v2(**kw)
    keys += [o["Key"] for o in r.get("Contents", [])]
    tok = r.get("NextContinuationToken")
    if not r.get("IsTruncated"): break

fixes = {}          # ключ маркера (без запятой) -> (новые координаты, комментарий, автор)
n_com = 0
for k in keys:
    try: cs = json.loads(s3.get_object(Bucket=BUCKET, Key=k)["Body"].read())
    except Exception: continue
    mkey = k.replace("corrections/", "").replace(".json", "")
    for c in cs:
        com = (c.get("comment") or "").strip()
        if not com: continue
        n_com += 1
        if c.get("verdict") == "correct": continue      # «верно» — координаты менять не надо
        m = COORD.search(com)
        if not m: continue
        la, lo = float(m.group(1)), float(m.group(2))
        if not valid(la, lo): continue
        fixes[mkey] = ((round(la, 6), round(lo, 6)), com, c.get("user", "anon"))

print(f"правок с комментарием: {n_com} | из них с явными координатами: {len(fixes)}")

ledger = json.load(open(LEDGER, encoding="utf-8")) if LEDGER.exists() else {}
new = {k: v for k, v in fixes.items() if k not in ledger}
print(f"новых к применению: {len(new)}")
if not new:
    print("нечего применять"); sys.exit(0)

# ---------- ищем строки таблицы по координате маркера ----------
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
H = [c.value for c in ws[1]]; C = {n: H.index(n) for n in H}

def mkey_of(la, lo):
    return re.sub(r"[^0-9A-Za-zА-Яа-я\-_.]", "", f"{round(float(la),6)},{round(float(lo),6)}")

rows_by_key = collections.defaultdict(list)
for row in ws.iter_rows(min_row=2):
    la, lo = row[C["lat"]].value, row[C["lon"]].value
    if la in (None, "", "ND") or str(la) == "None": continue
    try: rows_by_key[mkey_of(la, lo)].append(row)
    except Exception: pass

applied, missed = 0, []
for mkey, ((la, lo), com, who) in new.items():
    rows = rows_by_key.get(mkey)
    if not rows:
        missed.append(mkey); continue
    name = str(rows[0][C["nearest_locality"]].value or "?")
    old = (rows[0][C["lat"]].value, rows[0][C["lon"]].value)
    d = math.hypot((la - float(old[0])) * 111320,
                   (lo - float(old[1])) * 111320 * math.cos(math.radians(la))) / 1000
    print(f"  {name[:26]:26} {old[0]},{old[1]} -> {la},{lo}   (сдвиг {d:.0f} км)")
    print(f"      причина: {com[:100]}")
    if not DRY:
        for row in rows:
            row[C["lat"]].value = la
            row[C["lon"]].value = lo
            row[C["loc_confidence"]].value = "ok"      # координату подтвердил человек
        ledger[mkey] = {"locality": name, "to": [la, lo], "by": who, "comment": com[:200]}
    applied += 1

if missed:
    print(f"\nне нашлось строк для {len(missed)} правок (объект мог быть перемещён ранее)")
if DRY:
    print("\nсухой прогон — таблица не изменена"); sys.exit(0)

shutil.copy(XLSX, XLSX.with_suffix(".before_corrections.xlsx"))
wb.save(XLSX)
json.dump(ledger, open(LEDGER, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"\nприменено правок: {applied} -> {XLSX.name} (бэкап .before_corrections.xlsx)")

# ---------- отвечаем проверяющему в той же карточке ----------
# Пишем и по старой координате (там лежит его замечание — виден итог обсуждения),
# и по новой (туда объект переехал — там ответ увидят те, кто откроет его теперь).
VERIFY_API = os.environ.get("VERIFY_API", "https://functions.yandexcloud.net/d4evhi01uh5faspidt74")
import urllib.request
def say(section_id, text):
    body = {"section_id": section_id, "field": "location", "verdict": "correct",
            "comment": text, "user": "система"}
    req = urllib.request.Request(VERIFY_API, data=json.dumps(body).encode(), method="POST",
                                 headers={"Content-Type": "application/json"})
    try:
        urllib.request.urlopen(req, timeout=60).read(); return True
    except Exception as e:
        print("   не удалось отметить:", str(e)[:70]); return False

marked = 0
for mkey, ((la, lo), com, who) in new.items():
    info = ledger.get(mkey) or {}
    name = info.get("locality", "объект")
    text = f"ИСПРАВЛЕНО: координаты обновлены на {la}, {lo} по замечанию проверяющего"
    if say(f"{la},{lo}", text): marked += 1        # новая точка объекта
    say(mkey, text)                                 # и старая — чтобы замечание не осталось без ответа
print(f"отмечено как исправленное: {marked}")
print("дальше: досчитать высоты/радиусы, перепривязать страницы, пересобрать карту")
