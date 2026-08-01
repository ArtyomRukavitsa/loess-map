# Чистка неправдоподобных значений мощности (замечание геолога п.7 по сути: в поле «мощность»
# смешиваются разные величины). Проверка данных показала, что туда попадают ещё и посторонние числа:
# «Ипатово 0–2024 м» — это ГОД, «Лихвин 0.5–10000 м» — десять километров лёсса.
#
# Правила отсева (по одному числу, а не по всей строке — иначе теряется годное значение):
#   • похоже на год (1900–2030) — почти наверняка не мощность;
#   • изученная/видимая мощность > 300 м — таких разрезов в природе единицы, а в наших данных сотни;
#   • глубина скважины > 1000 м — уже не четвертичная толща (керны Востока и Байкала — не наш предмет).
# Если после отсева не осталось ничего — поле становится ND.
#
#   DRY=1 python 71_clean_thickness.py   — только показать, что уйдёт
import os, re, sys, json, shutil, pathlib, collections
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")
HERE = pathlib.Path(__file__).parent
XLSX = HERE / "records_clean_geo_v2.xlsx"
DRY = os.environ.get("DRY", "0") == "1"
MAX_TH = float(os.environ.get("MAX_TH", 300))       # изученная/видимая мощность
MAX_BORE = float(os.environ.get("MAX_BORE", 1000))  # глубина скважины

COLS = {"thickness_studied": MAX_TH, "thickness_visible": MAX_TH,
        "thickness_borehole_depth": MAX_BORE, "thickness_unspecified": MAX_TH}
NUM = re.compile(r"\d+(?:[.,]\d+)?")

def looks_year(x):
    return 1900 <= x <= 2030 and float(x).is_integer()

def clean(val, limit):
    """Возвращает (новое значение, что выброшено). Разбираем по числам, сохраняя разделители."""
    s = str(val or "").strip()
    if s in ("", "ND", "None"): return s, []
    dropped = []
    def repl(m):
        x = float(m.group(0).replace(",", "."))
        if looks_year(x) or x > limit:
            dropped.append(m.group(0)); return "\x00"
        return m.group(0)
    out = NUM.sub(repl, s)
    if not dropped: return s, []
    # убираем осиротевшие разделители и пустые куски диапазонов
    parts = [p.strip(" -–—;,") for p in out.split(";")]
    keep = []
    for p in parts:
        p = re.sub(r"\x00", "", p).strip(" -–—")
        p = re.sub(r"\s*[-–—]\s*$", "", p).strip()
        p = re.sub(r"^\s*[-–—]\s*", "", p).strip()
        if p and NUM.search(p): keep.append(p)
    return ("; ".join(keep) if keep else "ND"), dropped

wb = openpyxl.load_workbook(XLSX); ws = wb.active
H = [c.value for c in ws[1]]; C = {n: H.index(n) for n in H}

stat = collections.Counter(); examples = []
for row in ws.iter_rows(min_row=2):
    loc = str(row[C["nearest_locality"]].value or "")[:22]
    for col, limit in COLS.items():
        cell = row[C[col]]
        new, dropped = clean(cell.value, limit)
        if not dropped: continue
        stat["values"] += len(dropped); stat["cells"] += 1
        if new == "ND": stat["emptied"] += 1
        if len(examples) < 15:
            examples.append((loc, col.replace("thickness_", ""), str(cell.value)[:22], new, dropped[:3]))
        if not DRY: cell.value = new

print(f"ячеек с посторонними числами: {stat['cells']} | чисел выброшено: {stat['values']} | "
      f"полей опустело: {stat['emptied']}")
print()
for loc, col, old, new, dr in examples:
    print(f"   {loc:22} {col:16} «{old}» -> «{new}»   убрано: {', '.join(dr)}")

if DRY:
    print("\nсухой прогон — таблица не изменена"); sys.exit(0)
shutil.copy(XLSX, XLSX.with_suffix(".before_thickness.xlsx"))
wb.save(XLSX)
print(f"\n-> {XLSX.name} обновлён (бэкап .before_thickness.xlsx)")
