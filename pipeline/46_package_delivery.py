# Упаковка поставки заказчику: к gold-таблице добавляет n_sources + confidence_tier,
# делит на полную выгрузку и «надёжное ядро», кладёт в delivery/.
# Вход: records_gold.xlsx (32 gold-кол) + records_clean_geo.xlsx (n_sources, поля). Выравнивание по порядку строк (ID).
import openpyxl, os, sys, shutil
sys.stdout.reconfigure(encoding="utf-8")
HERE = os.path.dirname(os.path.abspath(__file__))
DELIV = os.path.join(HERE, "delivery"); os.makedirs(DELIV, exist_ok=True)

gwb = openpyxl.load_workbook("records_gold.xlsx"); gws = gwb.active
grows = list(gws.values); GH = list(grows[0]); GC = {n: i for i, n in enumerate(GH)}
cwb = openpyxl.load_workbook("records_clean_geo.xlsx"); cws = cwb.active
crows = list(cws.values); CH = list(crows[0]); CC = {n: i for i, n in enumerate(CH)}
assert len(grows) == len(crows), f"рассинхрон строк: gold {len(grows)} vs clean_geo {len(crows)}"

def nd(v): return v in (None, "", "ND", "No", "None")
def tier(g, c):
    geo = not nd(g[GC["N"]])
    nsrc = c[CC["n_sources"]] or 1
    try: nsrc = int(nsrc)
    except Exception: nsrc = 1
    rich = (not nd(g[GC["Thickness, m"]])) and (not nd(g[GC["Dating method"]])) and (not nd(g[GC["Stratigraphic position"]]))
    has_data = (not nd(g[GC["Type of deposits"]])) and (not nd(g[GC["Stratigraphic position"]]))
    if geo and (nsrc >= 2 or rich): return "A", nsrc      # локализован + подтверждён/богат
    if geo: return "B", nsrc                               # локализован, одно-источниковый
    if has_data: return "C", nsrc                          # есть данные, нет координат
    return "D", nsrc                                       # разреженный

OUT_H = GH + ["n_sources", "confidence_tier"]
full = openpyxl.Workbook(); fw = full.active; fw.title = "loess_sections"; fw.append(OUT_H)
core = openpyxl.Workbook(); kw = core.active; kw.title = "core"; kw.append(OUT_H)
counts = {"A": 0, "B": 0, "C": 0, "D": 0}; ncore = 0
for gr, cr in zip(grows[1:], crows[1:]):
    t, nsrc = tier(gr, cr)
    counts[t] += 1
    outrow = list(gr) + [nsrc, t]
    fw.append(outrow)
    if t == "A": kw.append(outrow); ncore += 1
full.save(os.path.join(DELIV, "loess_sections_full.xlsx"))
core.save(os.path.join(DELIV, "loess_sections_core.xlsx"))

# бонус: зарубежные + карту-ссылку
if os.path.exists("records_international.xlsx"):
    shutil.copy("records_international.xlsx", os.path.join(DELIV, "loess_sections_international_bonus.xlsx"))

print(f"delivery/ собрана:")
print(f"  loess_sections_full.xlsx  : {len(grows)-1} разрезов (все)")
print(f"  loess_sections_core.xlsx  : {ncore} разрезов (tier A — локализ.+надёжн./богатые)")
print(f"  confidence_tier: A={counts['A']} B={counts['B']} C={counts['C']} D={counts['D']}")
print(f"  + loess_sections_international_bonus.xlsx")
