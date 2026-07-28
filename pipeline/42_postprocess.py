# Постобработка записей: фильтр шума/иностранных -> нормализация чисел/локаций -> консолидация по разрезам
# (локация+регион) -> рабочая чистая БД. Вход sample_records.xlsx. Выход records_clean.xlsx + records_international.xlsx.
import openpyxl, re, sys, collections
sys.stdout.reconfigure(encoding="utf-8")
ILL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def xl(v): return ILL.sub("", str(v)) if v is not None else ""

wb = openpyxl.load_workbook("sample_records.xlsx"); ws = wb.active
rows = list(ws.values); H = list(rows[0]); data = rows[1:]
C = {n: H.index(n) for n in H}

# --- нормализация числа -> (min, max) в метрах ---
def parse_num(s):
    s = str(s).strip().lower()
    if s in ("nd", "", "none", "evidence", "-"): return None, None
    is_cm = bool(re.search(r"\bсм\b|\bcm\b", s))
    over = bool(re.search(r"свыше|более|>", s))
    under = bool(re.search(r"\bдо\b|<", s))
    nums = [float(x) for x in re.findall(r"\d+[.,]?\d*", s.replace(",", ".")) if x not in (".", "")]
    if not nums: return None, None
    if is_cm: nums = [x / 100 for x in nums]
    lo, hi = min(nums), max(nums)
    if over: return lo, None
    if under: return None, hi
    return lo, hi

# --- нормализация локации ---
PREF = re.compile(r"^(г|с|пос|дер|д|х|ст|пгт|аул|село|город|станица|хутор|деревня|оз|р)\.?\s+", re.I)
def norm_loc(s):
    s = re.sub(r"\s+", " ", str(s).strip())
    s = PREF.sub("", s).strip(" .,-")
    return s
def norm_admin(s):
    s = str(s).strip()
    return "" if s in ("ND", "None", "") else re.sub(r"\s+", " ", s)
FOREIGN_KW = ["каракорум", "гимала", "тибет", "альп", "кордильер", "аппалач", "сахар", "гоби", "пиреней",
              "апеннин", "скандинав", "гренланд", "антаркт", "китай", "инди", "пакистан", "иран", "ирак",
              "монгол", "япон", "корея", "вьетнам", "египет", "марокко", "алжир", "аргентин", "бразил",
              "мексик", "канад", "аляск", "гавай", "африк", "тунис", "ливи", "эфиоп", "танзан", "малави",
              "сша", "америк", "плато лёсс", "лёссовое плато", "чехи", "словак", "венгр", "румын", "болгар"]
def is_foreign(loc):                                            # латиница ИЛИ кириллическое-иностранное -> отдельно
    l = loc.lower()
    return bool(re.search(r"[A-Za-z]", loc)) or any(w in l for w in FOREIGN_KW)
def split_multi(s):
    return [x.strip() for x in str(s).split(";") if x.strip() and x.strip() not in ("ND", "None")]

# --- разбор записей ---
dom = collections.defaultdict(lambda: {"dep": set(), "str": set(), "dat": set(), "th": [], "el": [],
                                       "src": set(), "ev": [], "n": 0, "loc": "", "adm": ""})
intl = []
for r in data:
    loc0 = str(r[C["nearest_locality"]]).strip()
    if loc0 in ("ND", "", "None"): continue
    loc = norm_loc(loc0); adm = norm_admin(r[C["administrative_unit"]])
    rec = {"loc": loc, "adm": adm, "dep": split_multi(r[C["type_of_deposits"]]),
           "str": split_multi(r[C["stratigraphic_position"]]), "dat": split_multi(r[C["dating_methods"]]),
           "th": parse_num(r[C["thickness_m"]]), "el": parse_num(r[C["absolute_elevation_m"]]),
           "src": str(r[C["source_file"]]), "ev": str(r[C["evidence_locality"]])[:150]}
    if is_foreign(loc):
        intl.append(rec); continue
    k = (loc.lower(), adm.lower())
    g = dom[k]; g["loc"] = loc; g["adm"] = adm; g["n"] += 1
    g["dep"] |= set(rec["dep"]); g["str"] |= set(rec["str"]); g["dat"] |= set(rec["dat"])
    if rec["th"][0] is not None: g["th"].append(rec["th"][0])
    if rec["th"][1] is not None: g["th"].append(rec["th"][1])
    if rec["el"][0] is not None: g["el"].append(rec["el"][0])
    if rec["el"][1] is not None: g["el"].append(rec["el"][1])
    g["src"].add(rec["src"])
    if rec["ev"] and len(g["ev"]) < 2: g["ev"].append(rec["ev"])

# --- запись рабочей чистой БД ---
def num(v): return round(v, 2) if v is not None else "ND"
STR_ORD = ["Lower Pleistocene", "Middle Pleistocene", "Upper Pleistocene", "Holocene"]
wb2 = openpyxl.Workbook(); w = wb2.active; w.title = "sections"
w.append(["nearest_locality", "administrative_unit", "thickness_min_m", "thickness_max_m",
          "elevation_min_m", "elevation_max_m", "type_of_deposits", "stratigraphic_position",
          "dating_methods", "n_records", "n_sources", "sources", "evidence"])
sections = sorted(dom.values(), key=lambda g: -g["n"])
for g in sections:
    w.append([xl(g["loc"]), xl(g["adm"]),
              num(min(g["th"]) if g["th"] else None), num(max(g["th"]) if g["th"] else None),
              num(min(g["el"]) if g["el"] else None), num(max(g["el"]) if g["el"] else None),
              "; ".join(sorted(g["dep"])), "; ".join(sorted(g["str"], key=lambda x: STR_ORD.index(x) if x in STR_ORD else 9)),
              "; ".join(sorted(g["dat"])), g["n"], len(g["src"]),
              xl(" | ".join(sorted(g["src"]))[:300]), xl(" || ".join(g["ev"]))])
wb2.save("records_clean.xlsx")

# --- иностранные отдельно ---
wb3 = openpyxl.Workbook(); wi = wb3.active; wi.title = "international"
wi.append(["nearest_locality", "administrative_unit", "type_of_deposits", "stratigraphic_position", "dating_methods", "source_file", "evidence"])
for r in intl:
    wi.append([xl(r["loc"]), xl(r["adm"]), "; ".join(r["dep"]), "; ".join(r["str"]), "; ".join(r["dat"]), xl(r["src"]), xl(r["ev"])])
wb3.save("records_international.xlsx")

print(f"сырых записей: {len(data)}")
print(f"иностранных (отдельно): {len(intl)} -> records_international.xlsx")
print(f"отеч. записей консолидировано в РАЗРЕЗЫ: {sum(g['n'] for g in sections)} -> {len(sections)} уник. разрезов -> records_clean.xlsx")
print(f"с координатным потенциалом (локация+регион). thickness заполнен у {sum(1 for g in sections if g['th'])}, elevation у {sum(1 for g in sections if g['el'])}")
