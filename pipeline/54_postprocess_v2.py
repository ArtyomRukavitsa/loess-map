# Постобработка v2: sample_records_v2.xlsx -> консолидация по разрезам (локация+регион) с новыми полями.
# Фильтр иностранного. Типы мощности сливаются раздельно (studied/visible/borehole/unspecified -> min-max).
# Выход: records_clean_v2.xlsx (+ records_international_v2.xlsx).
import openpyxl, re, sys, collections
sys.stdout.reconfigure(encoding="utf-8")
ILL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def xl(v): return ILL.sub("", str(v)) if v is not None else ""

wb = openpyxl.load_workbook("sample_records_v2.xlsx"); ws = wb.active
rows = list(ws.values); H = list(rows[0]); C = {n: H.index(n) for n in H}; data = rows[1:]

def parse_num(s):
    s = str(s).strip().lower()
    if s in ("nd", "", "none", "-"): return None, None
    is_cm = bool(re.search(r"\bсм\b|\bcm\b", s))
    over = bool(re.search(r"свыше|более|>", s)); under = bool(re.search(r"\bдо\b|<", s))
    nums = [float(x) for x in re.findall(r"\d+[.,]?\d*", s.replace(",", ".")) if x not in (".", "")]
    if not nums: return None, None
    if is_cm: nums = [x / 100 for x in nums]
    lo, hi = min(nums), max(nums)
    if over: return lo, None
    if under: return None, hi
    return lo, hi

def parse_joined(s):                           # "2.24; 0.77–0.99" -> общий (min,max)
    los, his = [], []
    for part in str(s).split(";"):
        lo, hi = parse_num(part)
        if lo is not None: los.append(lo)
        if hi is not None: his.append(hi)
    return (min(los) if los else None), (max(his) if his else None)

PREF = re.compile(r"^(г|с|пос|дер|д|х|ст|пгт|аул|село|город|станица|хутор|деревня|оз|р)\.?\s+", re.I)
def norm_loc(s):
    s = re.sub(r"\s+", " ", str(s).strip()); return PREF.sub("", s).strip(" .,-")
def norm_admin(s):
    s = str(s).strip(); return "" if s in ("ND", "None", "") else re.sub(r"\s+", " ", s)
FOREIGN_KW = ["каракорум", "гимала", "тибет", "альп", "кордильер", "аппалач", "сахар", "гоби", "пиреней",
              "апеннин", "скандинав", "гренланд", "антаркт", "китай", "инди", "пакистан", "иран", "ирак",
              "монгол", "япон", "корея", "вьетнам", "египет", "марокко", "алжир", "аргентин", "бразил",
              "мексик", "канад", "аляск", "гавай", "африк", "тунис", "ливи", "эфиоп", "танзан", "малави",
              "сша", "америк", "чехи", "словак", "венгр", "румын", "болгар", "lake", "sea"]
def is_foreign(loc):
    l = loc.lower()
    return bool(re.search(r"[A-Za-z]", loc)) or any(w in l for w in FOREIGN_KW)
def split_multi(s):
    return [x.strip() for x in str(s).split(";") if x.strip() and x.strip() not in ("ND", "None")]

dom = collections.defaultdict(lambda: {"exc": set(), "geo": [], "dep": set(), "raw": set(), "str": set(),
    "dat": set(), "src": set(), "srck": set(), "ev": [], "n": 0, "loc": "", "adm": "",
    "th": {"studied": [], "visible": [], "borehole_depth": [], "unspecified": []}, "el": []})
intl = []
for r in data:
    loc0 = str(r[C["nearest_locality"]]).strip()
    if loc0 in ("ND", "", "None"): continue
    loc = norm_loc(loc0); adm = norm_admin(r[C["administrative_unit"]])
    if is_foreign(loc):
        intl.append(r); continue
    k = (loc.lower(), adm.lower())
    o = dom[k]; o["loc"] = o["loc"] or loc; o["adm"] = o["adm"] or adm; o["n"] += 1
    for cat in split_multi(r[C["type_of_deposits"]]): o["dep"].add(cat)
    for t in split_multi(r[C["deposit_raw_terms"]]): o["raw"].add(t)
    for s in split_multi(r[C["stratigraphic_position"]]): o["str"].add(s)
    for dm in split_multi(r[C["dating_methods"]]): o["dat"].add(dm)
    exc = str(r[C["excavation_type"]]).strip()
    if exc not in ("ND", "", "unspecified", "None"): o["exc"].add(exc)
    geo = str(r[C["geomorphic_position"]]).strip()
    if geo not in ("ND", "", "None") and geo not in o["geo"] and len(o["geo"]) < 5: o["geo"].append(geo)
    sk = str(r[C["source_kind"]]).strip()
    if sk not in ("ND", "", "None"): o["srck"].add(sk)
    for kind in o["th"]:
        o["th"][kind].append(str(r[C["thickness_" + kind]]))
    o["el"].append(str(r[C["absolute_elevation_m"]]))
    o["src"].add(str(r[C["source_file"]]))
    evl = str(r[C["evidence_locality"]]).strip()
    if evl and evl != "ND" and evl not in o["ev"] and len(o["ev"]) < 4: o["ev"].append(evl)

def rng(lo, hi):
    if lo is None and hi is None: return "ND"
    if lo is not None and hi is not None: return f"{lo:g}" if lo == hi else f"{lo:g}–{hi:g}"
    return f"≥{lo:g}" if lo is not None else f"≤{hi:g}"

OUT = ["nearest_locality", "administrative_unit", "excavation_type", "geomorphic_position",
       "type_of_deposits", "deposit_raw_terms", "thickness_studied", "thickness_visible",
       "thickness_borehole_depth", "thickness_unspecified", "elevation_min_m", "elevation_max_m",
       "stratigraphic_position", "dating_methods", "source_kinds", "n_records", "n_sources", "sources", "evidence"]
wb2 = openpyxl.Workbook(); w = wb2.active; w.title = "sections"; w.append(OUT)
n_th = n_el = 0
for (lk, ak), o in dom.items():
    ths = {kind: parse_joined("; ".join(o["th"][kind])) for kind in o["th"]}
    el_lo = min([parse_num(x)[0] for x in o["el"] if parse_num(x)[0] is not None], default=None)
    el_hi = max([parse_num(x)[1] for x in o["el"] if parse_num(x)[1] is not None], default=None)
    if any(ths[k] != (None, None) for k in ths): n_th += 1
    if el_lo is not None: n_el += 1
    w.append([xl(v) for v in [
        o["loc"], o["adm"], "; ".join(sorted(o["exc"])) or "ND", "; ".join(o["geo"]) or "ND",
        "; ".join(sorted(o["dep"])) or "ND", "; ".join(sorted(o["raw"])) or "ND",
        rng(*ths["studied"]), rng(*ths["visible"]), rng(*ths["borehole_depth"]), rng(*ths["unspecified"]),
        el_lo if el_lo is not None else "ND", el_hi if el_hi is not None else "ND",
        "; ".join(sorted(o["str"])) or "ND", "; ".join(sorted(o["dat"])) or "ND",
        "; ".join(sorted(o["srck"])) or "ND", o["n"], len(o["src"]), " | ".join(sorted(o["src"])),
        " || ".join(o["ev"]) or "ND"]])
wb2.save("records_clean_v2.xlsx")

wi = openpyxl.Workbook(); wiw = wi.active; wiw.title = "international"; wiw.append(H)
for r in intl: wiw.append([xl(v) for v in r])
wi.save("records_international_v2.xlsx")

print(f"сырых записей: {len(data)} | иностранных: {len(intl)} -> records_international_v2.xlsx")
print(f"отеч. консолидировано в РАЗРЕЗЫ: {len(dom)} -> records_clean_v2.xlsx")
print(f"  с мощностью (люб.тип): {n_th} | с высотой: {n_el}")
